import pandas as pd
from xlrd import open_workbook
import openpyxl
#pip install xlrd==1.2.0 
# The lastest version of xlrd is only support .xls file, so you can install the older version
def get_cedulas():
    wb= openpyxl.load_workbook('./Captcha/cedula.xlsx')
    da = wb['Hoja1']
    a = da['A']
    sizelist = len(a)
    cedulas = []
    for i in range(sizelist-1):
        iletra = 'A'+str(i+2)
        cedulas.append(da[iletra].value)
    wb.close()
    return cedulas

def get_ruc():
    wb= openpyxl.load_workbook('./Captcha/ruc.xlsx')
    da = wb['Hoja1']
    a = da['A']
    sizelist = len(a)
    rucs = []
    for i in range(sizelist-1):
        iletra = 'A'+str(i+2)
        rucs.append(da[iletra].value)
    wb.close()
    return rucs

def write_cedulas(salida, i):
    wb= openpyxl.load_workbook('./Captcha/cedula.xlsx')
    da = wb['Hoja1']
    iletra = 'C'+str(i+2)
    da[iletra] = salida
    wb.save('./Captcha/cedula.xlsx')
    return "Archivo Guardado"
    